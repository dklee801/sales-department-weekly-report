#!/usr/bin/env python3
"""
표준 양식 호환 Excel 보고서 생성기 (병합된 셀 처리 개선 버전)
2025년도 주간보고 양식_2.xlsx 구조에 맞춘 데이터 통합

주요 개선사항:
1. 매출 데이터 통합: '매출집계 데이터(raw)' 시트 하나에 월별/주차별 데이터 통합
2. 매출채권 표준화: 3개 시트 구조 (매출채권요약, 90일채권현황, 결제기간초과채권현황)
3. 표준 양식 테이블 구조 준수
4. 병합된 셀 처리 로직 개선
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
import shutil
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import warnings

# 새 구조에 맞는 import 수정
try:
    # 리팩토링된 구조에서 import 시도
    from ..data.processors.receivables_data_copier import ReceivablesDataCopier
except ImportError:
    try:
        # 백업: 상대 import 시도
        from ...modules.receivables_data_copier import ReceivablesDataCopier
    except ImportError:
        try:
            # 백업: 절대 import 시도
            from modules.receivables_data_copier import ReceivablesDataCopier
        except ImportError:
            # 최종 백업: 매출채권 복사 기능 비활성화
            ReceivablesDataCopier = None
            print("ReceivablesDataCopier를 찾을 수 없습니다. 매출채권 복사 기능이 비활성화됩니다.")

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class StandardFormatReportGenerator:
    """표준 양식 호환 보고서 생성기 (병합된 셀 처리 개선)"""
    
    def __init__(self, config_manager=None):
        self.logger = logging.getLogger('StandardFormatReportGenerator')
        self.config = config_manager
        self.result_path = None
        
        # 기본 경로들
        self.base_dir = Path(__file__).parent.parent.parent
        self.template_file = self.base_dir / "2025년도 주간보고 양식_2.xlsx"
        self.processed_dir = self.base_dir / "data/processed"
        self.report_dir = self.base_dir / "data/report"
        
        # 결과 파일 경로 설정
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        self.result_path = self.report_dir / f"주간보고서_{timestamp}.xlsx"
        
        # 표준 구조 매핑
        self.standard_structure = {
            'sales_raw_sheet': '매출집계 데이터(raw)',
            'receivables_summary_sheet': '매출채권요약',
            'ninety_days_sheet': '90일채권현황',
            'overdue_sheet': '결제기간초과채권현황',
            
            # 매출집계 데이터(raw) 시트 위치 매핑
            'monthly_start_pos': (2, 0),  # A3 (0-based: row=2, col=0)
            'weekly_start_pos': (2, 7),   # H3 (0-based: row=2, col=7)
            
            # 표준 헤더
            'monthly_headers': ['연도', 'month', '구동기', '일반부품', '무역', '티케이'],
            'weekly_headers': ['기간', '구동기', '일반부품', '무역', '티케이'],
        }
        
    def safe_clean(self, value):
        """데이터 정리 및 안전성 확보"""
        try:
            if value is None or pd.isna(value):
                return ""
            
            if isinstance(value, (int, float)):
                if pd.isna(value) or math.isinf(value) or math.isnan(value):
                    return 0
                # 너무 큰 값은 백만원 단위로 변환
                if abs(value) > 1e10:
                    return int(value / 1000000)
                return value
            
            if isinstance(value, str):
                clean_str = str(value).strip()
                if clean_str.lower() in ['nan', 'none', 'null', '']:
                    return ""
                return clean_str[:100]  # 길이 제한
            
            return str(value)[:100]
            
        except:
            return ""
    
    def safe_float(self, value):
        """값을 안전하게 float로 변환"""
        try:
            if value is None or pd.isna(value) or value == '':
                return 0.0
            return float(value)
        except:
            return 0.0

    def safe_write_cell(self, worksheet, row, col, value, skip_merged=True):
        """병합된 셀을 안전하게 처리하여 값 쓰기"""
        try:
            cell = worksheet.cell(row=row, column=col)
            
            # 병합된 셀 확인
            if isinstance(cell, MergedCell):
                if skip_merged:
                    self.logger.debug(f"병합된 셀 건너뜀: {row},{col}")
                    return False
                else:
                    # 병합 범위의 첫 번째 셀 찾기
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # 병합 범위의 첫 번째 셀에 값 쓰기
                            top_left = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left.value = value
                            self.logger.debug(f"병합된 셀의 상위 좌측에 값 쓰기: {merged_range.min_row},{merged_range.min_col} = {value}")
                            return True
            else:
                # 일반 셀에 값 쓰기
                cell.value = value
                return True
                
        except Exception as e:
            self.logger.error(f"셀 쓰기 실패 {row},{col}: {e}")
            return False

    def load_sales_data(self):
        """매출 데이터 로드 및 표준 구조로 변환"""
        try:
            sales_file = self.processed_dir / "매출집계_결과.xlsx"
            
            if not sales_file.exists():
                self.logger.warning("매출집계_결과.xlsx 파일이 없습니다.")
                return None, None
            
            # 월별 데이터 로드
            monthly_df = pd.read_excel(sales_file, sheet_name="월별", engine='openpyxl')
            weekly_df = pd.read_excel(sales_file, sheet_name="주차별", engine='openpyxl')
            
            self.logger.info(f"매출 데이터 로드 완료: 월별 {len(monthly_df)}행, 주차별 {len(weekly_df)}행")
            
            # 표준 구조로 변환
            monthly_standard = self.convert_monthly_to_standard(monthly_df)
            weekly_standard = self.convert_weekly_to_standard(weekly_df)
            
            return monthly_standard, weekly_standard
            
        except Exception as e:
            self.logger.error(f"매출 데이터 로드 실패: {e}")
            return None, None
    
    def convert_monthly_to_standard(self, df):
        """월별 데이터를 표준 구조로 변환"""
        try:
            if df.empty:
                return pd.DataFrame(columns=self.standard_structure['monthly_headers'])
            
            # 합계 행 제거
            df_clean = df[df['year'] != '합계'].copy()
            
            # 표준 컬럼 매핑
            standard_data = []
            
            for _, row in df_clean.iterrows():
                year = self.safe_clean(row.get('year', ''))
                month = self.safe_clean(row.get('month', ''))
                
                # 카테고리별 데이터 추출
                구동기 = self.safe_clean(row.get('구동기', 0))
                일반부품 = self.safe_clean(row.get('일반부품', 0))
                무역 = self.safe_clean(row.get('무역', 0))
                티케이 = self.safe_clean(row.get('티케이', 0))
                
                standard_data.append([year, month, 구동기, 일반부품, 무역, 티케이])
            
            result_df = pd.DataFrame(standard_data, columns=self.standard_structure['monthly_headers'])
            self.logger.info(f"월별 데이터 표준화 완료: {len(result_df)}행")
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"월별 데이터 변환 실패: {e}")
            return pd.DataFrame(columns=self.standard_structure['monthly_headers'])
    
    def convert_weekly_to_standard(self, df):
        """주차별 데이터를 표준 구조로 변환"""
        try:
            if df.empty:
                return pd.DataFrame(columns=self.standard_structure['weekly_headers'])
            
            # 합계 행 제거
            df_clean = df[df['기간'] != '합계'].copy()
            
            # 표준 컬럼 매핑
            standard_data = []
            
            for _, row in df_clean.iterrows():
                기간 = self.safe_clean(row.get('기간', ''))
                
                # 카테고리별 데이터 추출
                구동기 = self.safe_clean(row.get('구동기', 0))
                일반부품 = self.safe_clean(row.get('일반부품', 0))
                무역 = self.safe_clean(row.get('무역', 0))
                티케이 = self.safe_clean(row.get('티케이', 0))
                
                standard_data.append([기간, 구동기, 일반부품, 무역, 티케이])
            
            result_df = pd.DataFrame(standard_data, columns=self.standard_structure['weekly_headers'])
            self.logger.info(f"주차별 데이터 표준화 완료: {len(result_df)}행")
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"주차별 데이터 변환 실패: {e}")
            return pd.DataFrame(columns=self.standard_structure['weekly_headers'])
    
    def load_receivables_data(self):
        """매출채권 데이터 로드"""
        try:
            receivables_file = self.processed_dir / "채권_분석_결과.xlsx"
            
            if not receivables_file.exists():
                self.logger.warning("채권_분석_결과.xlsx 파일이 없습니다.")
                return None, None
            
            try:
                summary_df = pd.read_excel(receivables_file, sheet_name="요약", engine='openpyxl')
                calculation_df = pd.read_excel(receivables_file, sheet_name="계산 결과", engine='openpyxl')
                
                self.logger.info(f"매출채권 데이터 로드 성공: 요약 {summary_df.shape}, 계산 결과 {calculation_df.shape}")
                
                return summary_df, calculation_df
                
            except Exception as e:
                self.logger.error(f"매출채권 시트 읽기 실패: {e}")
                return None, None
                    
        except Exception as e:
            self.logger.error(f"매출채권 파일 접근 실패: {e}")
            return None, None
    
    def create_receivables_summary_table(self, summary_data, calculation_data):
        """매출채권요약 테이블 생성 (A1:I6 구조)"""
        try:
            # 표준 구조에 맞는 테이블 생성
            summary_table = []
            
            # 헤더 행들
            summary_table.append(['구분', '총 채권', '', '90일 초과', '', '총 채권 대비 90일', '', '90일 초과 채권', ''])
            summary_table.append(['', '', '', '채권(백만원)', '', '초과 채권율(%)', '', '증감(%p)', ''])
            summary_table.append(['', '전주', '금주', '전주', '금주', '전주', '금주', '전주', '금주'])
            
            # 데이터 행들 - 기본값으로 초기화
            dnd_row = ['DND', '', '', '', '', '', '', '', '']
            dni_row = ['DNI', '', '', '', '', '', '', '', '']
            total_row = ['합계', '', '', '', '', '', '', '', '']
            
            # 전주/금주 비교 데이터 추출 및 적용
            if summary_data is not None and calculation_data is not None:
                companies_data = self.extract_companies_receivables_data(summary_data, calculation_data)
                
                if companies_data:
                    # DND 데이터 적용
                    if 'DND' in companies_data:
                        dnd_data = companies_data['DND']
                        dnd_row[1] = dnd_data.get('total_receivables_prev', '')
                        dnd_row[2] = dnd_data.get('total_receivables_curr', '')
                        dnd_row[3] = dnd_data.get('over90_amount_prev', '')
                        dnd_row[4] = dnd_data.get('over90_amount_curr', '')
                        dnd_row[5] = dnd_data.get('over90_rate_prev', '')
                        dnd_row[6] = dnd_data.get('over90_rate_curr', '')
                        dnd_row[7] = dnd_data.get('over90_rate_change_prev', '')
                        dnd_row[8] = dnd_data.get('over90_rate_change', 0)
                    
                    # DNI 데이터 적용
                    if 'DNI' in companies_data:
                        dni_data = companies_data['DNI']
                        dni_row[1] = dni_data.get('total_receivables_prev', '')
                        dni_row[2] = dni_data.get('total_receivables_curr', '')
                        dni_row[3] = dni_data.get('over90_amount_prev', '')
                        dni_row[4] = dni_data.get('over90_amount_curr', '')
                        dni_row[5] = dni_data.get('over90_rate_prev', '')
                        dni_row[6] = dni_data.get('over90_rate_curr', '')
                        dni_row[7] = dni_data.get('over90_rate_change_prev', '')
                        dni_row[8] = dni_data.get('over90_rate_change', 0)
                    
                    # 합계 계산
                    total_row = self.calculate_receivables_total(dnd_row, dni_row)
            
            # 테이블 조립
            summary_table.append(dnd_row)
            summary_table.append(dni_row)
            summary_table.append(total_row)
            
            return pd.DataFrame(summary_table)
            
        except Exception as e:
            self.logger.error(f"매출채권요약 테이블 생성 실패: {e}")
            # 오류 시 기본 구조 반환
            return pd.DataFrame([
                ['구분', '총 채권', '', '90일 초과', '', '총 채권 대비 90일', '', '90일 초과 채권', ''],
                ['', '', '', '채권(백만원)', '', '초과 채권율(%)', '', '증감(%p)', ''],
                ['', '전주', '금주', '전주', '금주', '전주', '금주', '전주', '금주'],
                ['DND', '', '', '', '', '', '', '', ''],
                ['DNI', '', '', '', '', '', '', '', ''],
                ['합계', '', '', '', '', '', '', '', '']
            ])

    def calculate_receivables_total(self, dnd_row, dni_row):
        """매출채권요약 합계 행 계산 (가중평균)"""
        try:
            total_row = ['합계', '', '', '', '', '', '', '', '']
            
            # 기본값 추출
            dnd_total_prev = dnd_row[1] if dnd_row[1] != '' else 0
            dnd_total_curr = dnd_row[2] if dnd_row[2] != '' else 0
            dni_total_prev = dni_row[1] if dni_row[1] != '' else 0
            dni_total_curr = dni_row[2] if dni_row[2] != '' else 0
            
            dnd_over90_prev = dnd_row[3] if dnd_row[3] != '' else 0
            dnd_over90_curr = dnd_row[4] if dnd_row[4] != '' else 0
            dni_over90_prev = dni_row[3] if dni_row[3] != '' else 0
            dni_over90_curr = dni_row[4] if dni_row[4] != '' else 0
            
            dnd_rate_prev = dnd_row[5] if dnd_row[5] != '' else 0
            dnd_rate_curr = dnd_row[6] if dnd_row[6] != '' else 0
            dni_rate_prev = dni_row[5] if dni_row[5] != '' else 0
            dni_rate_curr = dni_row[6] if dni_row[6] != '' else 0
            
            # 1. 총채권 합계 (단순 합계)
            total_row[1] = int(dnd_total_prev) + int(dni_total_prev) if dnd_total_prev != '' and dni_total_prev != '' else ''
            total_row[2] = int(dnd_total_curr) + int(dni_total_curr) if dnd_total_curr != '' and dni_total_curr != '' else ''
            
            # 2. 90일초과 채권 합계 (단순 합계)
            total_row[3] = int(dnd_over90_prev) + int(dni_over90_prev) if dnd_over90_prev != '' and dni_over90_prev != '' else ''
            total_row[4] = int(dnd_over90_curr) + int(dni_over90_curr) if dnd_over90_curr != '' and dni_over90_curr != '' else ''
            
            # 3. 90일초과 비율 가중평균 계산
            if dnd_total_prev > 0 and dni_total_prev > 0 and dnd_rate_prev != '' and dni_rate_prev != '':
                weighted_avg_prev = ((dnd_total_prev * dnd_rate_prev) + (dni_total_prev * dni_rate_prev)) / (dnd_total_prev + dni_total_prev)
                total_row[5] = round(weighted_avg_prev, 2)
            else:
                total_row[5] = ''
            
            if dnd_total_curr > 0 and dni_total_curr > 0 and dnd_rate_curr != '' and dni_rate_curr != '':
                weighted_avg_curr = ((dnd_total_curr * dnd_rate_curr) + (dni_total_curr * dni_rate_curr)) / (dnd_total_curr + dni_total_curr)
                total_row[6] = round(weighted_avg_curr, 2)
            else:
                total_row[6] = ''
            
            # 4. H열, I열 증감 계산
            dnd_h = dnd_row[7] if len(dnd_row) > 7 and dnd_row[7] != '' else None
            dni_h = dni_row[7] if len(dni_row) > 7 and dni_row[7] != '' else None
            
            if dnd_h is not None and dni_h is not None:
                if dnd_total_prev > 0 and dni_total_prev > 0:
                    weighted_h = ((dnd_total_prev * float(dnd_h)) + (dni_total_prev * float(dni_h))) / (dnd_total_prev + dni_total_prev)
                    total_row[7] = round(weighted_h, 2)
                else:
                    total_row[7] = ''
            else:
                total_row[7] = ''
            
            dnd_i = dnd_row[8] if len(dnd_row) > 8 and dnd_row[8] != '' else None
            dni_i = dni_row[8] if len(dni_row) > 8 and dni_row[8] != '' else None
            
            if dnd_i is not None and dni_i is not None:
                if dnd_total_curr > 0 and dni_total_curr > 0:
                    weighted_i = ((dnd_total_curr * float(dnd_i)) + (dni_total_curr * float(dni_i))) / (dnd_total_curr + dni_total_curr)
                    total_row[8] = round(weighted_i, 2)
                else:
                    total_row[8] = ''
            else:
                total_row[8] = ''
            
            return total_row
            
        except Exception as e:
            self.logger.error(f"합계 행 계산 실패: {e}")
            return ['합계', '', '', '', '', '', '', '', '']

    def extract_companies_receivables_data(self, summary_data, calculation_data):
        """채권 데이터에서 DND/DNI 데이터 추출 (디버깅 강화 버전)"""
        try:
            companies_data = {}
            
            # 디버깅을 위한 로그 추가
            self.logger.info(f"=== 매출채권 데이터 추출 시작 ===")
            self.logger.info(f"summary_data 상태: {summary_data is not None}, shape: {summary_data.shape if summary_data is not None else 'None'}")
            self.logger.info(f"calculation_data 상태: {calculation_data is not None}, shape: {calculation_data.shape if calculation_data is not None else 'None'}")
            
            if summary_data is None or summary_data.empty:
                self.logger.warning("summary_data가 비어있거나 None입니다")
                return companies_data
            
            if calculation_data is None or calculation_data.empty:
                self.logger.warning("calculation_data가 비어있거나 None입니다")
                return companies_data
            
            # 실제 데이터 확인
            self.logger.info(f"calculation_data 첫 번째 행: {calculation_data.iloc[0].tolist() if len(calculation_data) > 0 else 'Empty'}")
            self.logger.info(f"calculation_data 컬럼명: {calculation_data.columns.tolist()}")
            
            # 전체 데이터 미리보기
            for i, row in calculation_data.iterrows():
                if i < 5:  # 첫 5개 행만 출력
                    self.logger.info(f"Row {i}: {row.tolist()}")
            
            # 계산 결과에서 전주/금주 데이터 추출
            processed_count = 0
            for idx, row in calculation_data.iterrows():
                try:
                    company_name_calc = str(row.iloc[0]).strip()
                    self.logger.info(f"처리 중인 회사명: '{company_name_calc}'")
                
                    if company_name_calc == '디앤드디':
                        company_key = 'DND'
                    elif company_name_calc == '디앤아이':
                        company_key = 'DNI'
                    elif company_name_calc == '합계':
                        company_key = '합계'
                    else:
                        self.logger.info(f"미인식 회사명: '{company_name_calc}' - 건너뜀")
                        continue
                    
                    self.logger.info(f"인식된 회사: {company_key}")
                
                    # 계산 결과 데이터 추출
                    total_prev = self.safe_float(row.iloc[1] if len(row) > 1 else 0)
                    total_curr = self.safe_float(row.iloc[2] if len(row) > 2 else 0)
                    total_change = self.safe_float(row.iloc[3] if len(row) > 3 else 0)
                    
                    over90_prev = self.safe_float(row.iloc[4] if len(row) > 4 else 0)
                    over90_curr = self.safe_float(row.iloc[5] if len(row) > 5 else 0)
                    over90_change = self.safe_float(row.iloc[6] if len(row) > 6 else 0)
                    
                    self.logger.info(f"{company_key} 추출 데이터: total_prev={total_prev}, total_curr={total_curr}, over90_prev={over90_prev}, over90_curr={over90_curr}")
                
                    # 비율 계산
                    over90_rate_prev = round((over90_prev / total_prev) * 100, 2) if total_prev > 0 else 0
                    over90_rate_curr = round((over90_curr / total_curr) * 100, 2) if total_curr > 0 else 0
                    
                    # 증감 계산
                    over90_rate_change = round(over90_rate_curr - over90_rate_prev, 2)
                    
                    companies_data[company_key] = {
                        'total_receivables_prev': int(total_prev / 1000000),  # 백만원 단위
                        'total_receivables_curr': int(total_curr / 1000000),  # 백만원 단위
                        'over90_amount_prev': int(over90_prev / 1000000),  # 백만원 단위
                        'over90_amount_curr': int(over90_curr / 1000000),  # 백만원 단위
                        'over90_rate_prev': over90_rate_prev,
                        'over90_rate_curr': over90_rate_curr,
                        'over90_rate_change': over90_rate_change,
                        'over90_rate_change_prev': "",  # 전전주 데이터 부족으로 빈 값
                    }
                    
                    self.logger.info(f"{company_key} 최종 데이터: {companies_data[company_key]}")
                    processed_count += 1
                    
                except Exception as e:
                    self.logger.error(f"행 {idx} 처리 중 오류: {e}")
                    continue
            
            self.logger.info(f"=== 매출채권 데이터 추출 완료: {processed_count}개 회사 처리됨 ===")
            self.logger.info(f"추출된 회사 목록: {list(companies_data.keys())}")
            
            return companies_data
            
        except Exception as e:
            self.logger.error(f"채권 데이터 추출 실패: {e}")
            return {}

    def generate_report(self, base_month=None, start_date_range=None):
        """표준 양식 호환 보고서 생성"""
        try:
            self.logger.info("=== 표준 양식 호환 보고서 생성 시작 ===")
            
            # 디렉토리 생성
            self.report_dir.mkdir(parents=True, exist_ok=True)
            
            # 템플릿 파일 확인
            if not self.template_file.exists():
                self.logger.error(f"템플릿 파일이 없습니다: {self.template_file}")
                return False
            
            # 템플릿 복사
            shutil.copy2(self.template_file, self.result_path)
            self.logger.info(f"템플릿 파일 복사 완료: {self.result_path}")
            
            # 데이터 로드
            monthly_data, weekly_data = self.load_sales_data()
            summary_data, calculation_data = self.load_receivables_data()
            
            # Excel 워크북 열기
            wb = load_workbook(str(self.result_path))
            
            # B1, D1 셀 설정 (옵셔널)
            if base_month or start_date_range:
                self.set_report_headers(wb, base_month, start_date_range)
            
            # 매출집계 데이터(raw) 시트 작성
            if monthly_data is not None or weekly_data is not None:
                self.write_sales_raw_sheet_safe(wb, monthly_data, weekly_data)
            
            # 매출채권요약 시트 작성 (비활성화 - receivables_data_copier에서 처리)
            # if summary_data is not None or calculation_data is not None:
            #     self.write_receivables_summary_sheet_safe(wb, summary_data, calculation_data)
            
            # 저장
            wb.save(str(self.result_path))
            wb.close()
            
            # 매출채권 데이터 자동 복사 추가
            self.logger.info("=== 매출채권 데이터 자동 복사 시작 ===")
            try:
                if ReceivablesDataCopier is not None:
                    copier = ReceivablesDataCopier()
                    copy_success = copier.copy_receivables_to_template(str(self.result_path))
                    
                    if copy_success:
                        self.logger.info("✅ 매출채권 데이터 자동 복사 완료")
                    else:
                        self.logger.warning("⚠️ 매출채권 데이터 복사 실패, 하지만 매출 데이터는 정상 생성")
                else:
                    self.logger.warning("⚠️ ReceivablesDataCopier를 사용할 수 없습니다")
                    
            except Exception as e:
                self.logger.error(f"매출채권 데이터 복사 중 오류: {e}")
                self.logger.warning("⚠️ 매출채권 데이터 없이 보고서 생성 완료")
            
            self.logger.info("=== 표준 양식 호환 보고서 생성 완료 ===")
            return True
            
        except Exception as e:
            self.logger.error(f"보고서 생성 실패: {e}")
            return False
    
    def write_sales_raw_sheet_safe(self, workbook, monthly_data, weekly_data):
        """매출집계 데이터(raw) 시트 안전 작성"""
        try:
            sheet_name = self.standard_structure['sales_raw_sheet']
            ws = workbook[sheet_name]
            
            if monthly_data is not None:
                # 월별 데이터를 A3부터 작성
                start_row = 3  # Excel 1-based
                start_col = 1  # Excel 1-based
                
                for row_idx, row_data in monthly_data.iterrows():
                    for col_idx, value in enumerate(row_data):
                        self.safe_write_cell(ws, start_row + row_idx, start_col + col_idx, value)
                
                self.logger.info(f"월별 데이터 작성 완료: {len(monthly_data)}행")
            
            if weekly_data is not None:
                # 주차별 데이터를 H3부터 작성
                start_row = 3  # Excel 1-based
                start_col = 8  # Excel 1-based (H열)
                
                for row_idx, row_data in weekly_data.iterrows():
                    for col_idx, value in enumerate(row_data):
                        self.safe_write_cell(ws, start_row + row_idx, start_col + col_idx, value)
                
                self.logger.info(f"주차별 데이터 작성 완료: {len(weekly_data)}행")
                
        except Exception as e:
            self.logger.error(f"매출집계 데이터(raw) 시트 작성 실패: {e}")
    
    def write_receivables_summary_sheet_safe(self, workbook, summary_data, calculation_data):
        """매출채권요약 시트 안전 작성 (병합된 셀 처리)"""
        try:
            sheet_name = self.standard_structure['receivables_summary_sheet']
            ws = workbook[sheet_name]
            
            # 매출채권요약 테이블 생성
            summary_table = self.create_receivables_summary_table(summary_data, calculation_data)
            
            # A1부터 테이블 작성 (병합된 셀 처리)
            start_row = 1  # Excel 1-based
            start_col = 1  # Excel 1-based
            
            for row_idx, row_data in summary_table.iterrows():
                for col_idx, value in enumerate(row_data):
                    # 병합된 셀이라도 강제로 값 쓰기 시도
                    self.safe_write_cell(ws, start_row + row_idx, start_col + col_idx, value, skip_merged=False)
            
            self.logger.info(f"매출채권요약 시트 작성 완료: {summary_table.shape}")
            
        except Exception as e:
            self.logger.error(f"매출채권요약 시트 작성 실패: {e}")
    
    def set_report_headers(self, workbook, base_month=None, start_date_range=None):
        """보고서 헤더 셀 설정 (B1: 기준월, D1: 시작주간 범위)"""
        try:
            # '1페이지' 시트 찾기
            if '1페이지' in workbook.sheetnames:
                ws = workbook['1페이지']
                
                # B1 셀에 기준월 설정
                if base_month:
                    ws['B1'] = base_month
                    self.logger.info(f"✅ B1 셀 설정 완료: {base_month}")
                
                # D1 셀에 시작주간 범위 설정
                if start_date_range:
                    ws['D1'] = start_date_range
                    self.logger.info(f"✅ D1 셀 설정 완료: {start_date_range}")
                    
            else:
                self.logger.warning("⚠️ '1페이지' 시트를 찾을 수 없습니다")
                
        except Exception as e:
            self.logger.error(f"보고서 헤더 설정 실패: {e}")

    def get_result_path(self):
        """결과 파일 경로 반환"""
        return str(self.result_path) if self.result_path else None


# 호환성을 위한 별칭들
XMLSafeReportGenerator = StandardFormatReportGenerator
WeeklyReportGenerator = StandardFormatReportGenerator


def generate_report():
    """함수 기반 호출을 위한 래퍼"""
    try:
        generator = StandardFormatReportGenerator()
        return generator.generate_report()
    except Exception as e:
        print(f"❌ 표준 양식 호환 보고서 생성 실패: {e}")
        return False


if __name__ == "__main__":
    # 테스트 실행
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    print("=== 표준 양식 호환 보고서 생성기 테스트 (병합된 셀 처리 개선) ===")
    
    generator = StandardFormatReportGenerator()
    success = generator.generate_report()
    
    if success:
        print(f"🎉 표준 양식 호환 보고서 생성 성공: {generator.get_result_path()}")
    else:
        print("💥 표준 양식 호완 보고서 생성 실패")
